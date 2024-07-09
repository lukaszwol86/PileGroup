import numpy as np
import pandas as pd
import math
import openpyxl


class PileGroup():
    np.set_printoptions(formatter={'float': '{: 0.2f}'.format})
    def pile_parameter(self,Ep,Gp,Ap,Jp,m,soil_type,**kwarg):
        self.Ep = Ep
        self.Gp = Gp
        self.Ap = Ap
        self.Jp = Jp
        self.m = m
        self.soil_type = soil_type
        if soil_type=="cohesive":
            self.kd= kwarg["kd"]
        elif soil_type=="friction":
            self.nh = kwarg["nh"]


    def loads(self,input_loads):
        self.Loads = pd.read_csv(input_loads, delimiter=";")
        self.R = []
        for index, row in self.Loads.iterrows():
            r = np.array([row.Fx * 1000, row.Fy * 1000, row.Fz * 1000, row.Mx * 1000, row.My * 1000, row.Mz * 1000])
            self.R.append(r)
    def loads_from_excel(self,file_name, sheet_name, cells_range, load_comb):
        df = pd.read_excel(file_name, sheet_name=sheet_name, usecols=cells_range, skiprows=1, engine='openpyxl')
        df.dropna(inplace=True)
        self.Loads = df.loc[df['Type'] == load_comb]
        self.R = []
        for index, row in self.Loads.iterrows():
            r = np.array([row.Fx * 1000, row.Fy * 1000, row.Fz * 1000, row.Mx * 1000, row.My * 1000, row.Mz * 1000])
            self.R.append(r)

    def single_load(self,input_loads):
        self.Loads = pd.read_csv(input_loads, delimiter=";")
        self.R = []
        row = self.Loads.loc[52]
        r = np.array([row.Fx * 1000, row.Fy * 1000, row.Fz * 1000, row.Mx * 1000, row.My * 1000, row.Mz * 1000])
        self.R.append(r)

    def input_piles(self,input_piles_pos):
        self.piles = pd.read_csv(input_piles_pos, delimiter=";")

    def input_pile_geom_excel(self, file_name, sheet_name, cells_range):
        df_piles = pd.read_excel(file_name, sheet_name=sheet_name, usecols=cells_range, skiprows=1, engine='openpyxl')
        df_piles.dropna(inplace=True)
        self.piles = df_piles



    def import_piles(self,piles_df):
        self.piles = piles_df[piles_df['Acti']==1]


    def piles_matrix(self):
        self.Kp = []
        for index, row in self.piles.iterrows():
            L = row.L
            k = np.zeros((6, 6))
            if self.soil_type == "no":
                k[0, 0] = self.m * 3 * self.Ep * self.Jp / L ** 3
                k[1, 1] = k[0, 0]
                k[2, 2] = self.Ap * self.Ep / L
                k[3, 3] = self.m * 3 * self.Ep * self.Jp / L
                k[4, 4] = k[3, 3]
                k[5, 5] = self.m * self.Gp * self.Jp / L
            elif self.soil_type == "cohesive":
                Le = (4 * self.Ep * self.Jp / (self.kd*row.redu)) ** (1 / 4)
                self.Le = Le
                k[0, 0] = (self.m + 1) * 2 * self.Ep * self.Jp / Le ** 3
                k[1, 1] = k[0, 0]
                k[2, 2] = self.Ap * self.Ep / L
                k[3, 3] = self.m * 2 * self.Ep * self.Jp / Le
                k[4, 4] = k[3, 3]
                k[5, 5] = self.m * self.Gp * self.Jp / L
                k[0, 4] = self.m * 2 * self.Ep * self.Jp / Le ** 2
                k[4, 0] = k[0, 4]
                k[1, 3] = -k[0, 4]
                k[3, 1] = -k[0, 4]
                # print(f"Pile {index+1} Le= {Le: 0.2f}")

            elif self.soil_type == "friction":
                Li = 1.8 * (self.Ep * self.Jp / (self.nh*row.redu)) ** (1 / 5)
                self.Li = Li
                k[0, 0] = (3 * self.m + 1) * 3 * self.Ep * self.Jp / Li ** 3
                k[1, 1] = k[0, 0]
                k[2, 2] = self.Ap * self.Ep / L
                k[3, 3] = self.m * 4 * self.Ep * self.Jp / Li
                k[4, 4] = k[3, 3]
                k[5, 5] = self.m * self.Gp * self.Jp / L
                k[0, 4] = self.m * 6 * self.Ep * self.Jp / Li ** 2
                k[4, 0] = k[0, 4]
                k[1, 3] = -k[0, 4]
                k[3, 1] = -k[0, 4]

            self.Kp.append(k)

        self.Dp = []
        self.Cp = []
        self.Api = []
        self.S = np.zeros((6, 6))
        i=0
        for index, row in self.piles.iterrows():
            cp = np.array([[1, 0, 0, 0, 0, 0],
                           [0, 1, 0, 0, 0, 0],
                           [0, 0, 1, 0, 0, 0],
                           [0, -row.z, row.y, self.m, 0, 0],
                           [row.z, 0, -row.x, 0, self.m, 0],
                           [-row.y, row.x, 0, 0, 0, self.m]])
            beta = math.atan(1 / row.N)
            alfa = math.radians(row.Angle)

            ap = np.array([[math.cos(beta) * math.cos(alfa), -math.sin(alfa), math.sin(beta) * math.cos(alfa), 0, 0, 0],
                           [math.cos(beta) * math.sin(alfa), math.cos(alfa), math.sin(beta) * math.sin(alfa), 0, 0, 0],
                           [-math.sin(beta), 0, math.cos(beta), 0, 0, 0],
                           [0, 0, 0, math.cos(beta) * math.cos(alfa), -math.sin(alfa), math.sin(beta) * math.cos(alfa)],
                           [0, 0, 0, math.cos(beta) * math.sin(alfa), math.cos(alfa), math.sin(beta) * math.sin(alfa)],
                           [0, 0, 0, -math.sin(beta), 0, math.cos(beta)]])

            dp = np.matmul(cp, ap)
            sp = np.matmul(np.matmul(dp, self.Kp[i]), dp.T)

            self.Dp.append(dp)
            self.S = np.add(self.S, sp)

            self.Api.append(ap)
            self.Cp.append(cp)
            i+=1

    def forces(self):
        self.df_pile_reactions = pd.DataFrame(columns=['LC', 'Pile', 'Vx', 'Vy', 'N', 'Mx', 'My', 'Mz'])
        self.df_pile_deformations = pd.DataFrame(columns=['LC', 'Pile', 'ux', 'uy', 'uz', 'psi_x', 'psi_y', 'psi_z'])

        self.U = []
        self.DispP = np.zeros((len(self.Loads), len(self.piles), 6))
        self.ForcsP = np.zeros((len(self.Loads), len(self.piles), 6))
        for i, reaction in enumerate(self.R):
            u = np.matmul(np.linalg.inv(self.S), reaction)
            self.U.append(u)
            for n, dp in enumerate(self.Dp):
                disp = np.matmul(dp.T, u)
                force = np.matmul(self.Kp[n], disp) / 1000
                # print(disp)
                pile_force = {'LC': i+1 , 'Pile': n+1 , 'Vx':force[0], 'Vy':force[1], 'N':force[2], 'Mx':force[3], 'My':force[4], 'Mz':force[5]}
                pile_displ = {'LC': i + 1, 'Pile': n + 1, 'ux': disp[0], 'uy': disp[1], 'uz': disp[2], 'psi_x': disp[3],
                              'psi_y': disp[4], 'psi_z': disp[5]}
                self.df_pile_reactions = self.df_pile_reactions.append(pile_force, ignore_index=True)
                self.df_pile_deformations = self.df_pile_deformations.append(pile_displ, ignore_index=True)
                self.DispP[i, n] = disp
                self.ForcsP[i, n] = force


        if self.soil_type== "friction":
            self.df_pile_reactions['M_max'] = (self.df_pile_reactions['Vx'] ** 2 + self.df_pile_reactions[
                'Vy'] ** 2) ** 0.5 * self.Li * 0.43

        elif self.soil_type== "cohesive":
            self.df_pile_reactions['M_max'] = (self.df_pile_reactions['Vx'] ** 2 + self.df_pile_reactions[
                'Vy'] ** 2) ** 0.5*self.Le*0.32
        else:
            self.df_pile_reactions['M_max'] = (self.df_pile_reactions['Mx'] ** 2 + self.df_pile_reactions[
                'My'] ** 2) ** 0.5



    def reaction_summary(self):
        self.force_summary = pd.pivot_table(data=self.df_pile_reactions, index=['Pile'], columns=['LC'], values=['N'])
        self.df_pile_deformations['uxy'] = (self.df_pile_deformations['ux']**2 + self.df_pile_deformations['uy']**2)**0.5
        self.displ_summary = pd.pivot_table(data=self.df_pile_deformations, index=['Pile'], columns=['LC'], values=['uxy'])

    def export_to_excel(self, file_name, comb_name):
        self.reaction_summary()
        wb =  openpyxl.load_workbook(file_name)
        if f'{comb_name}_force_summary' in wb.sheetnames:
            ws = wb[f'{comb_name}_force_summary']
            wb.remove(ws)
        if f'{comb_name}_forces' in wb.sheetnames:
            ws = wb[f'{comb_name}_forces']
            wb.remove(ws)
        if f'{comb_name}_displ_summary' in wb.sheetnames:
            ws = wb[f'{comb_name}_displ_summary']
            wb.remove(ws)
        if f'{comb_name}_displacement' in wb.sheetnames:
            ws = wb[f'{comb_name}_displacement']
            wb.remove(ws)
        wb.save(file_name)
        wb.close()


        with pd.ExcelWriter(file_name, mode='a') as writer:
            self.force_summary.to_excel(writer, sheet_name=f'{comb_name}_force_summary')
            self.df_pile_reactions.to_excel(writer, sheet_name=f'{comb_name}_forces')
            self.displ_summary.to_excel(writer, sheet_name=f'{comb_name}_displ_summary')
            self.df_pile_deformations.to_excel(writer, sheet_name=f'{comb_name}_displacement')




    def single_normal_forces(self):
        N_P_LC = np.zeros(len(self.piles))

        for pile in range(len(self.piles)):
                N_P_LC[ pile] = self.ForcsP[0, pile, 2]

        return N_P_LC

    def single_deformation(self):
        deform = ((self.U[0][0]) ** 2 + (self.U[0][1]) ** 2) ** 0.5 * 1000
        return deform


    def normal_forces(self):
        N_P_LC = np.zeros((len(self.Loads), len(self.piles)))

        for pile in range(len(self.piles)):
            for lc in range(len(self.Loads)):
                N_P_LC[lc, pile] = self.ForcsP[lc, pile, 2]

        return N_P_LC

    def deformations(self):
        deform = np.zeros((len(self.Loads), 4))

        for lc in range(len(self.Loads)):
            deform[lc, 0] = self.U[lc][0]*1000
            deform[lc, 1] = self.U[lc][1]*1000
            deform[lc, 2] = self.U[lc][2]*1000
            deform[lc, 3] = ((self.U[lc][0])**2+(self.U[lc][1])**2)**0.5*1000
        return deform


    def export_to_cae(self,file_name,param_dict):
        start_data = """<Indata>\n<Id>\n<Program>ec701</Program>\n<Version>2.1.4</Version>\n</Id>"""
        load_start = "\n<Last>\n<Kraft>"
        load_end = "\n</Kraft>\n</Last>"


        loads = ''
        i = 0
        for index, row in self.Loads.iterrows():
            Fx = row['Fx']
            Fy = row['Fy']
            Fz = row['Fz']
            Mx = row['Mx']
            My = row['My']
            Mz = row['Mz']

            load_temp = f'\n<Idx{i:03d}>\n<Smxd>{Mx:.1f}</Smxd>\n<Smyd>{My:.1f}</Smyd>\n<Smzd>{Mz:.1f}</Smzd>\n<Spxd>{Fx:.1f}</Spxd>\n<Spyd>{Fy:.1f}</Spyd>\n<Spzd>{Fz:.1f}</Spzd>\n</Idx{i:03d}>'
            loads += load_temp
            i += 1
        nr_loads = i
        pile_start = '\n<Pale>'

        pile_end = '\n</Pale>'

        pile_pos = ''
        for index, row in self.piles.iterrows():

            alfa = row['Angle']
            lengd = row['L']
            lutnig = row['N']
            typ = int(row['Type']-1)
            xkrd = row['x']
            ykrd = row['y']
            zkrd = row['z']
            L = row['L']

            pile_pos_temp = f'\n<Idx{index:03d}>\n<alpha>{alfa:.1f}</alpha>\n<lengd>{lengd:.1f}</lengd>\n<lutning>{lutnig:.1f}</lutning>\n<typ>{typ}</typ>\n<xkrd>{xkrd:.2f}</xkrd>\n<ykrd>{ykrd:.2f}</ykrd>\n<zkrd>{zkrd:.2f}</zkrd>\n</Idx{index:03d}>'
            pile_pos += pile_pos_temp

        pile_total = f'\n<antal>{index + 1}</antal>'

        sbp = 0.9  # m space between piles
        skp = 0.3  # m space to edge of foundation
        pile_spacing = f'\n<avstand>\n<kant>{skp:.1f}</kant>\n<pale>{sbp:.1f}</pale>\n</avstand>'

        if self.soil_type == "cohesive":
            soil_prop = f'\n<mark>\n<Rad00>\n<GrTyp>K</GrTyp>\n<kd>{self.kd/1000:.1f}</kd>\n<nh>0</nh>\n</Rad00>\n</mark>'

        elif self.soil_type == "friction":
            soil_prop = f'\n<mark>\n<Rad00>\n<GrTyp>F</GrTyp>\n<kd>0</kd>\n<nh>{self.nh:.1f}</nh>\n</Rad00>\n</mark>'
        else:
            soil_prop = f'\n<mark>\n<Rad00>\n<GrTyp>I</GrTyp>\n<kd>0</kd>\n<nh>0</nh>\n</Rad00>\n</mark>'

        param = '\n<param>\n<antal>\n<lastkomb>0</lastkomb>\n</antal>\n<plan>N</plan>\n</param>'
        #'\n<antal>\n<lastkomb>0</lastkomb>\n</antal>'

        platta = f"\n<platta>\n<xhoger>{param_dict['xhoger']:.2f}</xhoger>\n<xvanster>{param_dict['xvanster']:.2f}</xvanster>\n<ynere>{param_dict['ynere']:.2f}</ynere>\n<yuppe>{param_dict['yuppe']:.2f}</yuppe>\n</platta>"


        snitkraft = f'\n<Snittkraft>\n<Antal>{nr_loads}</Antal>\n</Snittkraft>'
        tvarsnitt = f"\n<Tvarsnitt>\n<Rad00>\n<Emodul>{(self.Ep / 10 ** 9):.1f}</Emodul>\n<Grupp>G</Grupp>\n<Insp>N</Insp>\n<Ix>{(self.Jp * 10 ** (12 - 8)):.4f}E+08</Ix>\n<Kv>{param_dict['Kv']/1000:.3f}E+07</Kv>\n<Lengd>{L:.1f}</Lengd>\n<Lfri>0</Lfri>\n<area>{(self.Ap * 10 ** (6)):.2f}</area>\n<bx>{param_dict['bx']:.2f}</bx>\n<by>{param_dict['by']:.2f}</by>\n<gmodul>{(self.Gp / 10 ** 9):.1f}</gmodul>\n</Rad00>\n<antal>1</antal>\n</Tvarsnitt>"

        end_data = "\n</Indata>"
        text = load_start + loads + load_end + pile_start + pile_pos + pile_total  + pile_spacing + soil_prop \
                 +param +platta+ pile_end+ snitkraft + tvarsnitt + end_data
        final_text = start_data + text.replace('.',',')
        with open(file_name, "w") as file:
            file.write(final_text)
        print(f"File {file_name} is saved")








if __name__ == "__main__":
    Piles = PileGroup()
    Piles.pile_parameter(Ep=33*10**9,Gp=12*10**9,Ap=109378*10**-6,Jp=58791*10**(4-4*3),m=0,soil_type="cohesive",kd=4615*1000)
    Piles.loads('Loads.csv')
    Piles.input_piles('piles.csv')
    Piles.export_to_cae('Support_1_ULS_test.xml',param_dict)

    # Piles.piles_matrix()
    # Piles.forces()



    #print(Piles.normal_forces())








